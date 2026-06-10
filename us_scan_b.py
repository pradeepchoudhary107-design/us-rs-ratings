"""
US RS Scanner — Part B (Second half of symbols)
Merges with docs/data_a.json, saves final docs/data.json
Sends email with full results.
"""
import yfinance as yf
import pandas as pd
import requests
import json
import io
import os
import time
import smtplib
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

EMAIL_SENDER   = os.environ.get("EMAIL_SENDER", "")
EMAIL_PASSWORD = os.environ.get("EMAIL_PASSWORD", "")
EMAIL_TO       = os.environ.get("EMAIL_TO", "")

def get_us_symbols():
    print("Fetching US stock list...")
    all_symbols = set()
    try:
        urls = [
            "https://www.nasdaqtrader.com/dynamic/SymDir/nasdaqlisted.txt",
            "https://www.nasdaqtrader.com/dynamic/SymDir/otherlisted.txt",
        ]
        for url in urls:
            r = requests.get(url, timeout=15)
            r.raise_for_status()
            lines = r.text.strip().split("\n")
            is_nasdaq = "nasdaq" in url
            for line in lines[1:]:
                parts = line.split("|")
                if len(parts) < 2:
                    continue
                sym = parts[0].strip()
                if not sym or len(sym) > 5:
                    continue
                if any(c in sym for c in ["^",".","/","$","+"]):
                    continue
                etf_col = 6 if is_nasdaq else 4
                if len(parts) > etf_col and parts[etf_col].strip().upper() == "Y":
                    continue
                all_symbols.add(sym)
        print(f"  Total symbols: {len(all_symbols)}")
    except Exception as e:
        print(f"  NASDAQ fetch failed: {e} — using fallback")
        all_symbols = set([
            "AAPL","MSFT","NVDA","AMZN","GOOGL","META","TSLA","JPM","LLY","V",
            "UNH","XOM","MA","JNJ","PG","HD","AVGO","MRK","COST","ABBV","CVX",
            "CRM","BAC","NFLX","AMD","PEP","KO","TMO","ACN","ADBE","WMT","MCD",
            "ABT","CSCO","LIN","DHR","TXN","NKE","PM","NEE","ORCL","INTC","RTX",
            "UNP","HON","BMY","AMGN","QCOM","LOW","CAT","GS","BA","IBM","SBUX",
            "INTU","SPGI","BLK","AXP","DE","ISRG","ADI","MDLZ","GILD","VRTX",
            "REGN","TJX","SYK","CI","TMUS","AMAT","MMC","ZTS","MO","CB","BDX",
            "EOG","BSX","SO","DUK","ITW","AON","CME","WM","PNC","CL","FCX",
            "PANW","KLAC","LRCX","SNPS","CDNS","ORLY","ADP","CTAS","MNST","FTNT",
            "MELI","CRWD","SNOW","PLTR","UBER","ABNB","NET","DDOG","MDB","COIN",
        ])
    symbols = sorted(list(all_symbols))
    # PART B = second half
    mid    = len(symbols) // 2
    part_b = symbols[mid:]
    print(f"  Part B: {len(part_b)} symbols (of {len(symbols)} total)")
    return part_b

def calc_rs(sym, start, end):
    try:
        ticker = yf.Ticker(sym)
        hist   = ticker.history(start=start, end=end, auto_adjust=True)
        if hist.empty or len(hist) < 150:
            return None
        p = hist["Close"].squeeze()
        if len(p) < 63:
            return None
        c_now = float(p.iloc[-1])
        c_3m  = float(p.iloc[-63])
        c_6m  = float(p.iloc[-126]) if len(p) >= 126 else float(p.iloc[0])
        c_9m  = float(p.iloc[-189]) if len(p) >= 189 else float(p.iloc[0])
        c_12m = float(p.iloc[-252]) if len(p) >= 252 else float(p.iloc[0])
        q4 = (c_now - c_3m)  / c_3m
        q3 = (c_3m  - c_6m)  / c_6m
        q2 = (c_6m  - c_9m)  / c_9m
        q1 = (c_9m  - c_12m) / c_12m
        score     = 0.40*q4 + 0.20*q3 + 0.20*q2 + 0.20*q1
        high_52w  = float(p.iloc[-252:].max()) if len(p) >= 252 else float(p.max())
        ret_1m    = round(((c_now - float(p.iloc[-21])) / float(p.iloc[-21]))*100, 1) if len(p) >= 21 else 0
        ret_3m    = round(((c_now - c_3m) / c_3m)*100, 1)
        ret_12m   = round(((c_now - c_12m) / c_12m)*100, 1)
        from_high = round(((c_now - high_52w) / high_52w)*100, 1)
        try:
            info     = ticker.info
            mktcap   = round(info.get("marketCap", 0) / 1e6, 0) if info.get("marketCap") else None
            sector   = info.get("sector",   "N/A")
            industry = info.get("industry", "N/A")
        except Exception:
            mktcap   = None
            sector   = "N/A"
            industry = "N/A"
        return {
            "sym": sym, "score": score, "price": round(c_now, 2),
            "mktcap": mktcap, "sector": sector, "industry": industry,
            "r1m": ret_1m, "r3m": ret_3m, "r12m": ret_12m,
            "h52": round(high_52w, 2), "fh": from_high,
        }
    except Exception:
        return None

def build_excel(df, fname):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    cols = [
        "Rank","Symbol","RS Rating","Strength",
        "Sector","Industry","Mkt Cap ($M)",
        "Price ($)","1M Ret%","3M Ret%","12M Ret%",
        "52W High","From 52W High%"
    ]
    display = df.rename(columns={
        "rank":"Rank","sym":"Symbol","rs":"RS Rating","strength":"Strength",
        "sector":"Sector","industry":"Industry","mktcap":"Mkt Cap ($M)",
        "price":"Price ($)","r1m":"1M Ret%","r3m":"3M Ret%","r12m":"12M Ret%",
        "h52":"52W High","fh":"From 52W High%",
    })[cols]
    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        display.to_excel(writer, sheet_name="US RS Ratings", index=False)
        ws = writer.sheets["US RS Ratings"]
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="111111")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center")
        for row_idx in range(2, len(display) + 2):
            rs_val = ws.cell(row=row_idx, column=3).value
            if   rs_val >= 90: bg, fg = "0a2e4a", "7dd3fc"
            elif rs_val >= 80: bg, fg = "1e3a5f", "93c5fd"
            elif rs_val >= 60: bg, fg = "633806", "FAC775"
            else:              bg, fg = "791F1F", "F7C1C1"
            for col in range(1, len(cols) + 1):
                c = ws.cell(row=row_idx, column=col)
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(color=fg, size=10)
                c.alignment = Alignment(horizontal="center")
        widths = [6,10,10,13,20,25,14,10,10,10,12,12,16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
    print(f"  Excel saved: {fname}")

def send_email(excel_path, df, date_str):
    if not EMAIL_PASSWORD:
        print("  EMAIL_PASSWORD not set — skipping.")
        return
    exc_count = len(df[df["rs"] >= 90])
    str_count = len(df[(df["rs"] >= 80) & (df["rs"] < 90)])
    subject   = f"US Market RS Rating — {date_str} | {len(df)} stocks | {exc_count} Exceptional"
    top10_rows = ""
    for _, row in df.head(10).iterrows():
        if row["rs"] >= 90:   bg, fg = "#0a2e4a", "#7dd3fc"
        elif row["rs"] >= 80: bg, fg = "#1e3a5f", "#93c5fd"
        else:                 bg, fg = "#633806", "#FAC775"
        star  = "★" if row["rs"] >= 90 else "◆"
        c12   = "#1D9E75" if row["r12m"] >= 0 else "#E24B4A"
        mktcap_s = f"${int(row['mktcap']):,}M" if row.get("mktcap") else "N/A"
        top10_rows += (
            "<tr>"
            f'<td style="padding:8px 10px;border-bottom:1px solid #222">{int(row["rank"])}</td>'
            f'<td style="padding:8px 10px;border-bottom:1px solid #222;font-weight:600">{row["sym"]}</td>'
            f'<td style="padding:8px 10px;border-bottom:1px solid #222">'
            f'<span style="background:{bg};color:{fg};padding:2px 8px;border-radius:5px;font-weight:700">'
            f'{row["rs"]} {star}</span></td>'
            f'<td style="padding:8px 10px;border-bottom:1px solid #222;color:#aaa;font-size:11px">{row.get("sector","N/A")}</td>'
            f'<td style="padding:8px 10px;border-bottom:1px solid #222;color:#aaa">{mktcap_s}</td>'
            f'<td style="padding:8px 10px;border-bottom:1px solid #222;color:{c12}">{row["r12m"]}%</td>'
            "</tr>"
        )
    body_html = f"""
    <div style="font-family:-apple-system,sans-serif;background:#0a0a0a;padding:2rem;max-width:680px;margin:0 auto">
      <div style="background:#111;border-radius:12px;padding:1.5rem;margin-bottom:1rem">
        <h1 style="color:#fff;font-size:1.2rem;font-weight:500;margin:0 0 4px">US Market RS Rating — Weekly Report</h1>
        <p style="color:#555;font-size:13px;margin:0">{date_str} · Full US Market (Part A + Part B combined)</p>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:1rem">
        <div style="background:#111;border-radius:8px;padding:1rem;text-align:center">
          <div style="color:#555;font-size:11px;margin-bottom:4px">TOTAL SCANNED</div>
          <div style="color:#fff;font-size:24px;font-weight:500">{len(df)}</div>
        </div>
        <div style="background:#061929;border:1px solid #0a2e4a;border-radius:8px;padding:1rem;text-align:center">
          <div style="color:#38bdf8;font-size:11px;margin-bottom:4px">EXCEPTIONAL 90+</div>
          <div style="color:#7dd3fc;font-size:24px;font-weight:500">{exc_count}</div>
        </div>
        <div style="background:#0e1f07;border:1px solid #27500A;border-radius:8px;padding:1rem;text-align:center">
          <div style="color:#639922;font-size:11px;margin-bottom:4px">STRONG 80+</div>
          <div style="color:#C0DD97;font-size:24px;font-weight:500">{str_count}</div>
        </div>
      </div>
      <div style="background:#111;border-radius:12px;padding:1.5rem;margin-bottom:1rem">
        <h2 style="color:#aaa;font-size:13px;font-weight:500;margin:0 0 1rem;text-transform:uppercase">Top 10 US Stocks</h2>
        <table style="width:100%;border-collapse:collapse;font-size:13px">
          <tr style="color:#555">
            <th style="text-align:left;padding:6px 10px;font-weight:500">#</th>
            <th style="text-align:left;padding:6px 10px;font-weight:500">Symbol</th>
            <th style="text-align:left;padding:6px 10px;font-weight:500">RS</th>
            <th style="text-align:left;padding:6px 10px;font-weight:500">Sector</th>
            <th style="text-align:left;padding:6px 10px;font-weight:500">Mkt Cap</th>
            <th style="text-align:left;padding:6px 10px;font-weight:500">12M Ret</th>
          </tr>
          {top10_rows}
        </table>
      </div>
      <p style="color:#444;font-size:12px;text-align:center;margin:0">
        Full Excel attached · Auto every Saturday 10 AM IST
      </p>
    </div>"""
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = EMAIL_SENDER
    msg["To"]      = EMAIL_TO
    msg.attach(MIMEText(body_html, "html"))
    with open(excel_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(excel_path)}"')
    msg.attach(part)
    try:
        recipients = [e.strip() for e in EMAIL_TO.split(",")]
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, recipients, msg.as_string())
        print(f"  Email sent to: {recipients}")
    except Exception as e:
        print(f"  Email failed: {e}")

def main():
    t0       = time.time()
    date_str = datetime.now().strftime("%d %b %Y")
    print(f"US RS Scan PART B — {date_str}")
    symbols  = get_us_symbols()
    end      = datetime.today()
    start    = end - timedelta(days=420)
    results  = []
    for i, sym in enumerate(symbols):
        if (time.time() - t0) > 4800:
            print(f"  Time limit — stopping at {i}")
            break
        r = calc_rs(sym, start, end)
        if r:
            results.append(r)
        if (i + 1) % 100 == 0:
            elapsed = time.time() - t0
            rem     = (elapsed / (i + 1)) * (len(symbols) - i - 1) / 60
            print(f"  [{i+1}/{len(symbols)}] {len(results)} ok | ~{rem:.0f} min left")
        time.sleep(0.25)
    print(f"\n  Part B complete: {len(results)} stocks")

    # Load Part A results
    part_a_results = []
    try:
        with open("docs/data_a.json", "r") as f:
            data_a = json.load(f)
            part_a_results = data_a.get("stocks", [])
        print(f"  Part A loaded: {len(part_a_results)} stocks")
    except Exception as e:
        print(f"  Part A not found: {e} — using Part B only")

    # Combine A + B
    all_results = part_a_results + results
    print(f"  Combined total: {len(all_results)} stocks")

    if not all_results:
        print("No results at all.")
        return

    # Rank all combined stocks 1-99
    df       = pd.DataFrame(all_results)
    df["rs"] = df["score"].rank(pct=True).apply(
        lambda p: max(1, min(99, round(1 + p * 98)))
    ).astype(int)
    df       = df.sort_values("rs", ascending=False).drop(columns=["score"])
    df["rank"] = range(1, len(df) + 1)
    df["strength"] = df["rs"].apply(
        lambda r: "Exceptional" if r >= 90 else (
            "Strong" if r >= 80 else ("Average" if r >= 60 else "Weak")
        )
    )

    # Save final JSON for dashboard
    records = df.to_dict(orient="records")
    payload = {
        "updated":     datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "updated_ist": datetime.now().strftime("%d %b %Y %I:%M %p IST"),
        "total":       len(records),
        "stocks":      records,
    }
    os.makedirs("docs", exist_ok=True)
    with open("docs/data.json", "w") as f:
        json.dump(payload, f, separators=(",", ":"))
    print(f"  Final JSON saved: {len(records)} stocks")

    ts = datetime.now().strftime("%Y%m%d")
    xl = f"US_RS_{ts}.xlsx"
    build_excel(df, xl)
    send_email(xl, df, date_str)
    print(f"\nDone! {len(all_results)} stocks | {round((time.time()-t0)/60,1)} min")

if __name__ == "__main__":
    main()
